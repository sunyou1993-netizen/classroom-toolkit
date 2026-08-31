"""간단교육 퀴즈 문항집 엑셀을 만듭니다 — 게임마다 시트 하나.

  node scripts/make-xlsx.mjs && python3 scripts/make-xlsx.py

시트: 안내 · 속담 · 사자성어 · 환경 · 안전 · 학교폭력 · 교가
"""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

ROOT = os.getcwd()
자료 = json.load(open(os.path.join(ROOT, "scripts", "_xlsx-data.json"), encoding="utf-8"))
저장경로 = os.path.join(ROOT, "간단교육_퀴즈_문항집.xlsx")

# ── 보기 규칙 ────────────────────────────────────────────────
# 한글이 들어가므로 한글이 있는 글꼴을 씁니다.
# 맑은 고딕은 윈도우 기본, 맥에서는 애플고딕으로 자동 대체됩니다.
글꼴 = "맑은 고딕"
먹 = "1F2A26"
회 = "5C6A66"

게임색 = {
    "속담": "8A4B7C",
    "사자성어": "1F6F8B",
    "환경": "2C7A4B",
    "안전": "9A5E14",
    "학교폭력": "3A55A0",
    "교가": "5B6B66",
    "안내": "0F6455",
}

가는선 = Side(style="thin", color="DDE4E0")
테두리 = Border(left=가는선, right=가는선, top=가는선, bottom=가는선)


def 제목행(ws, 열이름들, 색):
    ws.append(열이름들)
    for c in range(1, len(열이름들) + 1):
        셀 = ws.cell(row=ws.max_row, column=c)
        셀.font = Font(name=글꼴, size=10, bold=True, color="FFFFFF")
        셀.fill = PatternFill("solid", fgColor=색)
        셀.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        셀.border = 테두리
    ws.row_dimensions[ws.max_row].height = 24


def 폭주기(ws, 폭들):
    for i, w in enumerate(폭들, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def 줄꾸미기(ws, 첫줄, 줄바꿈열, 가운데열=(), 링크열=None):
    """본문 줄에 글꼴·테두리·줄바꿈을 입힙니다."""
    for r in range(첫줄, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            셀 = ws.cell(row=r, column=c)
            셀.border = 테두리
            셀.font = Font(name=글꼴, size=10, color=먹)
            if c in 줄바꿈열:
                셀.alignment = Alignment(vertical="top", wrap_text=True)
            elif c in 가운데열:
                셀.alignment = Alignment(horizontal="center", vertical="top")
            else:
                셀.alignment = Alignment(vertical="top")
        if 링크열:
            셀 = ws.cell(row=r, column=링크열)
            주소 = 셀.value
            if isinstance(주소, str) and 주소.startswith("http"):
                셀.hyperlink = Hyperlink(ref=셀.coordinate, target=주소)
                셀.font = Font(name=글꼴, size=9, color="0F6455", underline="single")


def 시트틀(ws, 제목열수):
    """제목줄 고정 + 거르기 단추."""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(제목열수)}{ws.max_row}"


# ══════════════════════════════════════════════════════════════
wb = Workbook()

# ── 1. 안내 ───────────────────────────────────────────────────
ws = wb.active
ws.title = "안내"
ws.sheet_properties.tabColor = 게임색["안내"]
폭주기(ws, [34, 10, 14, 52])

ws["A1"] = "간단교육 퀴즈 문항집"
ws["A1"].font = Font(name=글꼴, size=18, bold=True, color=게임색["안내"])
ws["A2"] = f"교실 화면(수업도우미 · 간단교육 퀴즈)에 나가는 게임 6개의 전체 문항입니다.   만든 날 {자료['만든날']}"
ws["A2"].font = Font(name=글꼴, size=10, color=회)
ws.row_dimensions[1].height = 26

ws["A4"] = "게임별 문항 수"
ws["A4"].font = Font(name=글꼴, size=11, bold=True, color=먹)

ws.append([])  # 5행 비움
제목행(ws, ["게임", "문항 수", "형태", "출처"], 게임색["안내"])
첫줄 = ws.max_row + 1

# 문항 수는 각 시트를 직접 세는 수식으로 넣습니다(문항이 바뀌면 같이 바뀝니다).
줄들 = [
    ("속담", "=COUNT(속담!A:A)", "빈칸 채우기", "없음 — 앱에 원래 들어 있던 콘텐츠"),
    ("사자성어", "=COUNT(사자성어!A:A)", "빈칸 채우기", "없음 — 앱에 원래 들어 있던 콘텐츠"),
    ("환경", "=COUNT(환경!A:A)", "O/X 고르기", "기후에너지환경부 · 기상청 등 기관 공식 자료"),
    ("안전", "=COUNT(안전!A:A)", "O/X 고르기", "행정안전부 · 법제처 · 질병관리청 등 기관 공식 자료"),
    ("학교폭력", "=COUNT(학교폭력!A:A)", "O/X 고르기", "법제처 · 교육부 등 기관 공식 자료"),
    ("교가", 0, "학교마다 다름", "학교 교가 — 배포 학교가 직접 넣습니다"),
]
for 이름, 수, 형, 출 in 줄들:
    ws.append([이름, 수, 형, 출])
    ws.cell(row=ws.max_row, column=1).font = Font(name=글꼴, size=10, bold=True, color=게임색[이름])

합계줄 = ws.max_row + 1
ws.append(["합계", f"=SUM(B{첫줄}:B{ws.max_row})", "", ""])
for c in range(1, 5):
    셀 = ws.cell(row=합계줄, column=c)
    셀.font = Font(name=글꼴, size=10, bold=True, color=먹)
    셀.fill = PatternFill("solid", fgColor="EDF1EF")

for r in range(첫줄, 합계줄 + 1):
    for c in range(1, 5):
        셀 = ws.cell(row=r, column=c)
        셀.border = 테두리
        if not 셀.font.bold:
            셀.font = Font(name=글꼴, size=10, color=먹)
        셀.alignment = Alignment(
            horizontal="center" if c == 2 else "left", vertical="top", wrap_text=(c == 4))

# 기관 집계
표2 = 합계줄 + 3
ws.cell(row=표2 - 1, column=1, value="O/X 세 게임(환경 · 안전 · 학교폭력)의 근거 기관").font = Font(
    name=글꼴, size=11, bold=True, color=먹)
ws.cell(row=표2, column=1, value="기관")
ws.cell(row=표2, column=2, value="문항 수")
for c in (1, 2):
    셀 = ws.cell(row=표2, column=c)
    셀.font = Font(name=글꼴, size=10, bold=True, color="FFFFFF")
    셀.fill = PatternFill("solid", fgColor=게임색["안내"])
    셀.alignment = Alignment(horizontal="center", vertical="center")
    셀.border = 테두리
for i, (기관이름, 수) in enumerate(자료["기관"], start=1):
    r = 표2 + i
    ws.cell(row=r, column=1, value=기관이름).font = Font(name=글꼴, size=10, color=먹)
    ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
    ws.cell(row=r, column=2, value=수).font = Font(name=글꼴, size=10, color=먹)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    for c in (1, 2):
        ws.cell(row=r, column=c).border = 테두리
# 기관 표 합계 — 문항 수가 O/X 세 게임 합과 맞는지 스스로 확인합니다.
r합 = 표2 + len(자료["기관"]) + 1
ws.cell(row=r합, column=1, value="합계 (환경 + 안전 + 학교폭력)")
ws.cell(row=r합, column=2, value=f"=SUM(B{표2 + 1}:B{표2 + len(자료['기관'])})")
for c in (1, 2):
    셀 = ws.cell(row=r합, column=c)
    셀.font = Font(name=글꼴, size=10, bold=True, color=먹)
    셀.fill = PatternFill("solid", fgColor="EDF1EF")
    셀.border = 테두리
    셀.alignment = Alignment(horizontal="center" if c == 2 else "left", vertical="center")

# 읽는 사람이 알아야 할 것
메모줄 = 표2 + len(자료["기관"]) + 4
메모 = [
    ("이 문항들에 대해", True),
    ("· 환경 · 안전 · 학교폭력 문항은 위 기관의 공식 자료를 근거로 새로 썼습니다. 기관 문장을 그대로 옮긴 곳은 없습니다.", False),
    ("  공공누리 제1유형(상업적 이용 허용)으로 공개된 자료가 없어, 기준과 수치만 근거로 삼고 초등학생 눈높이로 다시 썼습니다.", False),
    ("· 2026년 8월에 전 문항과 인용 URL 전부를 다시 열어 대조했고, 사실이 틀린 문항 1건과 근거가 맞지 않는 문항 3건을 지우거나 고쳤습니다.", False),
    ("· 속담 · 사자성어는 퀴즈 앱에 원래 들어 있던 콘텐츠입니다. 기관 출처가 없으므로 근거로 내세울 수 없습니다.", False),
    ("  표기 · 뜻풀이를 검수해 18건을 고쳤고, 답이 문제 안에 그대로 보이던 빈칸 32곳을 다른 글자로 바꿨습니다.", False),
    ("· 교가는 학교마다 다르므로 정해진 문항이 없습니다. 학교 교가를 넣기 전에는 목록에 나오지 않습니다.", False),
    ("", False),
    ("다음에 확인할 것", True),
    ("· 학교폭력예방법이 2027년 1월 1일에 개정 시행됩니다. 2026년 12월 안에 학교폭력 문항의 조문 근거를 한 번 더 확인해 주세요.", False),
    ("· 실태조사를 근거로 한 문항은 해마다 새 조사가 나오면 확인이 필요합니다.", False),
]
for i, (글, 굵게) in enumerate(메모):
    r = 메모줄 + i
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    셀 = ws.cell(row=r, column=1, value=글)
    셀.font = Font(name=글꼴, size=10, bold=굵게, color=먹 if 굵게 else 회)
    셀.alignment = Alignment(vertical="center")

# ── 2. 속담 ───────────────────────────────────────────────────
ws = wb.create_sheet("속담")
ws.sheet_properties.tabColor = 게임색["속담"]
제목행(ws, ["번호", "카테고리", "속담", "빈칸1", "빈칸2", "뜻", "출처"], 게임색["속담"])
for x in 자료["속담"]:
    ws.append([x["번호"], x["카테고리"], x["속담"], x["빈칸1"], x["빈칸2"], x["뜻"], x["출처"]])
폭주기(ws, [6, 14, 38, 7, 7, 56, 30])
줄꾸미기(ws, 2, 줄바꿈열={3, 6, 7}, 가운데열={1, 4, 5})
시트틀(ws, 7)

# ── 3. 사자성어 ───────────────────────────────────────────────
ws = wb.create_sheet("사자성어")
ws.sheet_properties.tabColor = 게임색["사자성어"]
제목행(ws, ["번호", "카테고리", "사자성어", "한자", "빈칸1", "빈칸2", "뜻", "출처"], 게임색["사자성어"])
for x in 자료["사자성어"]:
    ws.append([x["번호"], x["카테고리"], x["사자성어"], x["한자"],
               x["빈칸1"], x["빈칸2"], x["뜻"], x["출처"]])
폭주기(ws, [6, 14, 12, 10, 7, 7, 60, 30])
줄꾸미기(ws, 2, 줄바꿈열={7, 8}, 가운데열={1, 3, 4, 5, 6})
시트틀(ws, 8)

# ── 4~6. O/X 세 게임 ─────────────────────────────────────────
for 이름 in ("환경", "안전", "학교폭력"):
    ws = wb.create_sheet(이름)
    ws.sheet_properties.tabColor = 게임색[이름]
    제목행(ws, ["번호", "카테고리", "문항", "정답", "해설", "근거기관", "근거문서", "출처URL"], 게임색[이름])
    for x in 자료[이름]:
        ws.append([x["번호"], x["카테고리"], x["문항"], x["정답"], x["해설"],
                   x["근거기관"], x["근거문서"], x["출처URL"]])
    폭주기(ws, [6, 16, 52, 6, 50, 26, 52, 40])
    줄꾸미기(ws, 2, 줄바꿈열={3, 5, 6, 7}, 가운데열={1, 4}, 링크열=8)
    for r in range(2, ws.max_row + 1):
        # 정답 O 는 초록, X 는 벽돌색으로
        셀 = ws.cell(row=r, column=4)
        셀.font = Font(name=글꼴, size=11, bold=True,
                       color="2C7A4B" if 셀.value == "O" else "A8442F")
        # 근거문서 이름을 눌러도 원문으로 가게 합니다(URL 칸을 안 봐도 되도록).
        문서 = ws.cell(row=r, column=7)
        주소 = ws.cell(row=r, column=8).value
        if isinstance(주소, str) and 주소.startswith("http"):
            문서.hyperlink = Hyperlink(ref=문서.coordinate, target=주소)
            문서.font = Font(name=글꼴, size=10, color="0F6455", underline="single")
    시트틀(ws, 8)
    # 종이로 뽑을 때는 URL 칸(H)을 뺍니다. 여덟 칸을 한 장에 우겨넣으면 글씨가 너무 작아집니다.
    # 파일 안에는 그대로 있고, 근거문서 이름이 링크라 화면에서는 눌러서 갈 수 있습니다.
    ws.print_area = f"A1:G{ws.max_row}"

# ── 7. 교가 ───────────────────────────────────────────────────
ws = wb.create_sheet("교가")
ws.sheet_properties.tabColor = 게임색["교가"]
폭주기(ws, [18, 90])
ws["A1"] = "교가"
ws["A1"].font = Font(name=글꼴, size=16, bold=True, color=게임색["교가"])
교 = 자료["교가"]
줄들 = [
    ("형태", "빈칸 채우기 — 우리 학교 교가의 빈칸을 채웁니다"),
    ("문항", "정해진 문항이 없습니다. 교가는 학교마다 다릅니다."),
    ("지금 상태",
     f"켜짐 · {교['학교']} · {교['절']}절 {교['줄']}줄" if 교["켜짐"]
     else "꺼짐 · 아직 학교 교가를 넣지 않아 퀴즈 목록에 나오지 않습니다"),
    ("출처", "학교 교가 — 배포하는 학교가 직접 넣습니다"),
    ("", ""),
    ("넣는 법", "1) quiz/scripts/school-song.json 에 학교 이름과 가사를 적습니다."),
    ("", "2) node scripts/set-school-song.mjs 를 실행합니다."),
    ("", "3) node scripts/make-sw.mjs 를 실행합니다."),
    ("", "대괄호 [ ] 로 감싼 두 글자가 빈칸이 됩니다. 대괄호를 안 쓰면 줄마다 알아서 고릅니다."),
    ("", "교가를 넣으면 퀴즈 목록에 교가 카드가 자동으로 켜집니다."),
    ("", ""),
    ("왜 꺼 두었나",
     "퀴즈 앱에는 특정 학교(서울신답초등학교)의 교가가 들어 있었습니다. "
     "그대로 배포하면 다른 학교 아이들에게 남의 학교 교가를 가르치게 되므로, "
     "학교 교가를 넣기 전까지는 목록에서 빼 두었습니다."),
]
r = 3
for 왼, 오 in 줄들:
    ws.cell(row=r, column=1, value=왼).font = Font(name=글꼴, size=10, bold=True, color=먹)
    ws.cell(row=r, column=1).alignment = Alignment(vertical="top")
    셀 = ws.cell(row=r, column=2, value=오)
    셀.font = Font(name=글꼴, size=10, color=먹 if 오 else 회)
    셀.alignment = Alignment(vertical="top", wrap_text=True)
    r += 1

# ── 인쇄 설정 ─────────────────────────────────────────────────
# 종이로 뽑을 때 옆으로 잘리지 않게, 그리고 장마다 제목줄이 다시 나오게 합니다.
for sh in wb.worksheets:
    sh.sheet_view.showGridLines = False
    sh.page_setup.orientation = "landscape"
    sh.page_setup.fitToWidth = 1
    sh.page_setup.fitToHeight = 0
    sh.sheet_properties.pageSetUpPr.fitToPage = True
    if sh.title not in ("안내", "교가"):
        sh.print_title_rows = "1:1"
wb.active = 0
wb.save(저장경로)
print("만듦:", 저장경로)
for s in wb.worksheets:
    print(f"  {s.title:<10} {s.max_row - 1 if s.title not in ('안내', '교가') else '-':>5} 줄")
